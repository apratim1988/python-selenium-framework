import openpyxl


#reading value from xcel
book = openpyxl.load_workbook("C:\\Users\\aprat\\PycharmProjects\\PythonSelFramework\\pythonxcel.xlsx")
sheet = book.active

Dict ={}
cell = sheet.cell(row=1,column=2)
print(cell.value)

#writing value from xcel
sheet.cell(row=2,column=2).value = "Apratim"

#to check t he number of rows/columns in the sheet
sheet.max_row
sheet.max_column

#to print the values of all row and all columns using for loop

for i in range(1,sheet.max_row+1):
    for j in range(1,sheet.max_column+1):
        print(sheet.cell(row=i,column=j))

#to print the values of a selected row
for i in range(1,sheet.max_row+1):
    if sheet.cell(row=i,column=1).value == "TestCase2":
        for j in range(1,sheet.max_column+1):
            print(sheet.cell(row=i,column=j))


#from the below code,the firstname ,lastname and gender will be printed
for i in range(1,sheet.max_row+1):
    if sheet.cell(row=i,column=1).value == "TestCase2":
        for j in range(1,sheet.max_column+1):
            Dict[sheet.cell(row=1,column=j).value] = sheet.cell(row=i,column=j).value
print(Dict)


